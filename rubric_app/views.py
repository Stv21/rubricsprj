from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Subject, Rubric, Marks, Student, Teacher, Classroom
from .forms import CustomUserCreationForm, RubricUploadForm, ClassroomForm, JoinClassroomForm
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import views as auth_views
from django.contrib import messages
from django.conf import settings
from django.urls import reverse
from django.http import HttpResponse
import xlwt
import uuid
import os
from gradio_client import Client, handle_file
import csv
from .forms import StudentProfileForm
import json
import base64
import requests
from .forms import ExcelUploadForm
from openpyxl import load_workbook
from django.contrib.auth import get_user_model

User = get_user_model()

from django.conf import settings

def custom_login(request, *args, **kwargs):
    if request.user.is_authenticated:
        return redirect('dashboard_redirect')
    else:
        return auth_views.LoginView.as_view(template_name='login.html')(request, *args, **kwargs)

# Dashboard redirect view to determine user type
@login_required
def dashboard_redirect(request):
    if request.user.is_superuser:
        return redirect('admin_dashboard')
    elif request.user.is_teacher:
        return redirect('teacher_dashboard')
    elif request.user.is_student:
        return redirect('student_dashboard')
    else:
        return redirect('login')

# Student dashboard view
@login_required
def student_dashboard(request):
    student = request.user.student
    classrooms = student.classrooms.all()
    join_form = JoinClassroomForm()

    # Assign a home URL based on the user role
    home_url = reverse('student_dashboard')

    # Assign light colors to classrooms
    light_colors = ["#f0f0f0", "#e0e0f0", "#d0d0f0", "#c0c0f0", "#b0b0f0"]
    for idx, classroom in enumerate(classrooms):
        classroom.color = light_colors[idx % len(light_colors)]

    if request.method == 'POST' and 'join_classroom' in request.POST:
        join_form = JoinClassroomForm(request.POST)
        if join_form.is_valid():
            code = join_form.cleaned_data['code']
            try:
                classroom = Classroom.objects.get(code=code)
                student.classrooms.add(classroom)
                messages.success(request, "Joined classroom successfully!")
            except Classroom.DoesNotExist:
                join_form.add_error('code', 'Invalid code')

    return render(
        request,
        'student_dashboard.html',
        {
            'classrooms': classrooms,
            'join_form': join_form,
            'home_url': home_url,
        },
    )
@login_required
def student_profile(request):
    student = get_object_or_404(Student, user=request.user)
    
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('student_profile')
    else:
        form = StudentProfileForm(instance=student)

    return render(request, 'student_profile.html', {
        'form': form,
        'student': student,
        'classrooms': student.classrooms.all()
    })
@login_required
def delete_rubric(request, rubric_id):
    rubric = get_object_or_404(Rubric, id=rubric_id, student=request.user.student)
    if request.method == 'POST':
        rubric.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

# Teacher dashboard view with classroom creation functionality
@login_required

def teacher_dashboard(request):
    teacher = request.user.teacher
    classrooms = Classroom.objects.filter(teacher=teacher)
    create_form = ClassroomForm()

    if request.method == 'POST':
        create_form = ClassroomForm(request.POST)
        if create_form.is_valid():
            classroom = create_form.save(commit=False)
            classroom.teacher = teacher
            classroom.code = uuid.uuid4().hex[:10]  
            classroom.save()
            messages.success(request, f"Classroom created! Code: {classroom.code}")
            return redirect('teacher_dashboard')
        else:
            messages.error(request, "Failed to create classroom. Please check the details.")

    return render(request, 'teacher_dashboard.html', {
        'classrooms': classrooms,
        'create_form': create_form
    })

    
# View classroom details for teacher, displaying enrolled students and their rubrics
@login_required
def get_enrolled_students(request, classroom_id):
    # Fetch the classroom, accessible to both students and teachers
    classroom = get_object_or_404(Classroom, id=classroom_id)
    
    # Ensure the user has access to the classroom
    if request.user.is_teacher and classroom.teacher.user != request.user:
        return JsonResponse({'error': 'You do not have access to this classroom'}, status=403)
    if request.user.is_student and not request.user.student.classrooms.filter(id=classroom_id).exists():
        return JsonResponse({'error': 'You are not enrolled in this classroom'}, status=403)
    
    # Fetch the list of enrolled students
    students = classroom.students.all().values_list('user__username', flat=True)
    return JsonResponse({'students': list(students)})


@login_required
def view_classroom(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id)
    context = {'classroom': classroom}

    if request.user.is_teacher:
        # Ensure the teacher owns the classroom
        teacher = request.user.teacher
        if classroom.teacher != teacher:
            messages.error(request, "You do not have permission to view this classroom.")
            return redirect("teacher_dashboard")

        # Fetch students and rubrics
        students = classroom.students.all()
        rubrics = Rubric.objects.filter(classroom=classroom)

        # Add to the context
        context["students"] = students
        context["rubrics"] = rubrics
        context["is_teacher"] = True

    elif request.user.is_student:
        # Ensure the student is enrolled in the classroom
        student = request.user.student
        if not student.classrooms.filter(id=classroom_id).exists():
            messages.error(request, "You are not part of this classroom.")
            return redirect("student_dashboard")

        # Fetch rubrics for the logged-in student
        rubrics = Rubric.objects.filter(classroom=classroom, student=student)

        # Add to the context
        context["rubrics"] = rubrics
        context["students"] = classroom.students.all()  # Allow students to view enrolled peers
        context["is_student"] = True
        
        

    else:
        messages.error(request, "Invalid user role.")
        return redirect("dashboard_redirect")

    return render(request, "view_classroom.html", context)


@login_required
def export_grades(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id)
    
    if not request.user.is_teacher or classroom.teacher.user != request.user:
        return HttpResponse('Unauthorized', status=401)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{classroom.name}_grades.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Grades')

    # Styles
    header_style = xlwt.XFStyle()
    header_style.font.bold = True
    
    subheader_style = xlwt.XFStyle()
    subheader_style.font.bold = True
    subheader_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
    
    data_style = xlwt.XFStyle()
    
    total_style = xlwt.XFStyle()
    total_style.font.bold = True

    # Write main headers
    ws.write(0, 0, "ROLL NO.", header_style)
    ws.write(0, 1, "STUDENT NAME", header_style)
    
    # Get all students in the classroom
    students = list(classroom.students.all())
    def roll_key(s):
        rn = s.roll_number
        # blank or non‐numeric rolls go last
        if not rn or not rn.isdigit():
            return (1, float('inf'))
        return (0, int(rn))
    students.sort(key=roll_key)
    
    # Setup lab headers (merge cells for lab headers)
    for lab_num in range(1, 11):
        col_start = 2 + (lab_num-1)*5  # Each lab has 5 columns
        ws.write_merge(0, 0, col_start, col_start+4, f"LAB {lab_num}", header_style)
        
        # Lab subheaders
        ws.write(1, col_start, "k (4)", subheader_style)
        ws.write(1, col_start+1, "Pe (5)", subheader_style)
        ws.write(1, col_start+2, "C (3)", subheader_style)
        ws.write(1, col_start+3, "Pu (3)", subheader_style)
        ws.write(1, col_start+4, "T", subheader_style)
    
    # Set column widths
    ws.col(0).width = 2500  # Roll No
    ws.col(1).width = 8000  # Student Name
    
    # Data rows
    for row_idx, student in enumerate(students, start=2):
        # Roll no.
        ws.write(row_idx, 0, student.roll_number or "", data_style)
        # Student name
        ws.write(row_idx, 1, student.full_name or "", data_style)

        # First fetch this student’s rubrics, then build the dict
        rubrics = Rubric.objects.filter(student=student, classroom=classroom)
        rubric_dict = {r.experiment_number: r for r in rubrics}

        for lab_num in range(1, 11):
            col = 2 + (lab_num - 1) * 5
            r = rubric_dict.get(lab_num)
            if r:
                ws.write(row_idx, col,     r.knowledge or 0, data_style)
                ws.write(row_idx, col+1, r.performance or 0, data_style)
                ws.write(row_idx, col+2, r.content_neatness or 0, data_style)
                ws.write(row_idx, col+3, r.punctuality_submission or 0, data_style)
                total = sum([
                    r.knowledge or 0, r.performance or 0,
                    r.content_neatness or 0, r.punctuality_submission or 0
                ])
                ws.write(row_idx, col+4, total, total_style)
            else:
                for i in range(5):
                    ws.write(row_idx, col+i, 0, data_style if i < 4 else total_style)

    wb.save(response)
    return response

# ---------------------- Rubric Processing ----------------------
def delete_classroom(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user.teacher)
    
    if request.method == 'POST':
        classroom.delete()  # Delete the classroom
        return redirect('teacher_dashboard')  # Redirect to the teacher dashboard after deletion

    return redirect('teacher_dashboard')  # Fallback
# View student rubrics for a specific classroom, accessible by the teacher
@login_required
def view_student_rubrics(request, classroom_id, student_id):
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user.teacher)
    student = get_object_or_404(Student, id=student_id, classrooms=classroom)
    rubrics = Rubric.objects.filter(student=student, classroom=classroom)
    return render(request, 'upload_rubrics.html', {
        'classroom': classroom,
        'student': student,
        'rubrics': rubrics
    })

# Admin dashboard (if applicable)
@login_required
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')

# Join classroom view
@login_required
def join_classroom(request):
    if request.method == 'POST':
        join_form = JoinClassroomForm(request.POST)
        if join_form.is_valid():
            code = join_form.cleaned_data['code']
            classroom = get_object_or_404(Classroom, code=code)
            student = request.user.student
            student.classrooms.add(classroom)
            messages.success(request, "Successfully joined the classroom!")
            return redirect('student_dashboard')
    return redirect('student_dashboard')
@login_required
def upload_rubric(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id, students=request.user.student)
    
    if request.method == 'POST':
        form = RubricUploadForm(request.POST, request.FILES)
        if form.is_valid():
            rubric_file = request.FILES['rubric_file']
            # save file locally
            file_path = os.path.join(settings.MEDIA_ROOT, 'rubrics', rubric_file.name)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'wb+') as dest:
                for chunk in rubric_file.chunks():
                    dest.write(chunk)

            # read & base64-encode
            with open(file_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode("utf8")

            payload = {"data": [f"data:image/jpeg;base64,{b64}"]}
            api_url = "https://xaviers--rubricease-infer.modal.run/api/predict"

            try:
                resp = requests.post(
                    api_url,
                    headers={"Content-Type": "application/json"},
                    data=json.dumps(payload),
                    timeout=60
                )
                resp.raise_for_status()
                j = resp.json()
                data = j.get("data", [])
                if len(data) >= 6:
                    _, know, pun, perf, cont, total = data[:6]
                    total = float(total)
                else:
                    raise ValueError(f"Unexpected response format: {j}")

                # save rubric record
                rubric = form.save(commit=False)
                rubric.classroom = classroom
                rubric.student = request.user.student
                rubric.knowledge = int(know)
                rubric.performance = int(perf)
                rubric.content_neatness = int(cont)
                rubric.punctuality_submission = int(pun)
                rubric.total_marks = total
                rubric.rubric_file.name = os.path.join('rubrics', rubric_file.name)
                rubric.save()
                Marks.objects.create(rubric=rubric)

                return JsonResponse({
                    'success': True,
                    'message': f'Rubric for Lab {rubric.experiment_number} processed successfully!',
                    'imageUrl': rubric.rubric_file.url,
                    'knowledge': know,
                    'performance': perf,
                    'content_neatness': cont,
                    'punctuality_submission': pun,
                    'total_marks': total,
                })

            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})

    else:
        form = RubricUploadForm()

    return render(request, 'upload_rubric.html', {
        'form': form,
        'classroom': classroom
    })

# ---------------------- Registration Views ----------------------

# Student registration
def student_register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_student = True
            user.save()
            Student.objects.create(user=user)
            login(request, user)
            return redirect('student_dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'student_register.html', {'form': form})

@login_required
def unenroll_from_classroom(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id)
    student = request.user.student
    
    if classroom.students.filter(id=student.id).exists():
        classroom.students.remove(student)  # Remove the student from the classroom
        return redirect('student_dashboard')  # Redirect back to the student dashboard after unenrollment

    return redirect('student_dashboard')  # Fallback

# Teacher registration
def teacher_register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_teacher = True
            user.save()
            Teacher.objects.create(user=user)
            login(request, user)
            return redirect('teacher_dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'teacher_register.html', {'form': form})

# --    -------------------- Admin Views ----------------------

@login_required
@user_passes_test(lambda u: u.is_superuser or u.is_teacher)
def bulk_create_students(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id)
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            wb = load_workbook(request.FILES['excel_file'])
            ws = wb.active
            created = 0

            # Expect header row: ["ROLL NO.", "STUDENT NAME"]
            for roll, name in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if not roll or not name:
                    continue

                # keep student full name
                full_name = str(name).lstrip("- ").strip()

                # ↓ NEW: roll.no as username & password 
                username = str(roll)
                password = str(roll)

                if User.objects.filter(username=username).exists():
                    continue

                # create the Django user…
                user = User.objects.create_user(username=username, password=password)
                user.is_student = True
                user.save()

                # then create the Student profile
                student = Student.objects.create(
                    user=user,
                    roll_number=str(roll),
                    full_name=full_name
                )
                student.classrooms.add(classroom)
                created += 1

            messages.success(request, f"{created} students created and enrolled.")
            return redirect('view_classroom', classroom_id=classroom.id)
    else:
        form = ExcelUploadForm()

    return render(request, "bulk_upload_students.html", {
        "form": form,
        "classroom": classroom
    })

from django.contrib import messages
from django.shortcuts import redirect

@login_required
@user_passes_test(lambda u: u.is_superuser or u.is_teacher)
def delete_classroom_student(request, classroom_id, student_id):
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user.teacher)
    student   = get_object_or_404(Student, id=student_id, classrooms=classroom)
    if request.method == "POST":
        # unenroll
        classroom.students.remove(student)
        # delete student profile and user account
        user = student.user
        student.delete()
        user.delete()
        messages.success(request, "Student login removed and account deleted.")
    return redirect('view_classroom', classroom_id=classroom_id)

@login_required
@user_passes_test(lambda u: u.is_superuser or u.is_teacher)
def delete_all_classroom_students(request, classroom_id):
    classroom = get_object_or_404(Classroom, id=classroom_id, teacher=request.user.teacher)
    if request.method == "POST":
        students = list(classroom.students.all())
        for student in students:
            # remove enrollment
            classroom.students.remove(student)
            # delete profile + user
            user = student.user
            student.delete()
            user.delete()
        messages.success(request, f"Deleted {len(students)} students and their accounts.")
    return redirect('view_classroom', classroom_id=classroom_id)

